import time, unicodedata

from openpyxl import load_workbook
from smartcard.System import readers
from smartcard.Exceptions import NoCardException

# Blacklist loading
TITLES = [
    "นาย", "นาง", "นางสาว", "เด็กชาย", "เด็กหญิง", "ด.ช.", "ด.ญ.",
    "mr.", "mrs.", "ms.", "miss"
]

def normalize(text: str) -> str:
    text = unicodedata.normalize("NFC", text)
    text = " ".join(text.split()).strip()
    text_lower = text.lower()
    for title in TITLES:
        if text_lower.startswith(title):
            text = text[len(title):].strip()
            break
    return text.lower()

def clean_id(text: str) -> str:
    return "".join(char for char in str(text or "") if char.isdigit())

def load_blacklist(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    id_records = {}
    name_records = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = clean_id(row[idx["ID CARD"]]) if "ID CARD" in idx else ""
        name = normalize(str(row[idx["Name(TH)"]] or "")) if "Name(TH)" in idx else ""
        status = str(row[idx["Status"]] or "").strip() if "Status" in idx else ""
        area = str(row[idx["Area"]] or "").strip() if "Area" in idx and row[idx["Area"]] is not None else ""
        caught_area = str(row[idx["พื้นที่/บริเวณที่ตรวจพบ"]] or "").strip() if "พื้นที่/บริเวณที่ตรวจพบ" in idx and row[idx["พื้นที่/บริเวณที่ตรวจพบ"]] is not None else ""
        occurance_date = str(row[idx["OccurenceDate"]] or "").strip() if "OccurenceDate" in idx and row[idx["OccurenceDate"]] is not None else ""
        company_name = str(row[idx["Company_Name"]] or "").strip() if "Company_Name" in idx and row[idx["Company_Name"]] is not None else ""
        department = str(row[idx["Dept."]] or "").strip() if "Dept." in idx and row[idx["Dept."]] is not None else ""
        card_number = str(row[idx["CARD No."]] or "").strip() if "CARD No." in idx and row[idx["CARD No."]] is not None else ""
        position = str(row[idx["Position"]] or "").strip() if "Position" in idx and row[idx["Position"]] is not None else ""
        affiliation = str(row[idx["สังกัด"]] or "").strip() if "สังกัด" in idx and row[idx["สังกัด"]] is not None else ""
        allegation = str(row[idx["Allegate (ข้อหา)"]] or "").strip() if "Allegate (ข้อหา)" in idx else ""
        start_date = str(row[idx["Start date"]] or "").strip() if "Start date" in idx and row[idx["Start date"]] is not None else ""
        due_date = str(row[idx["Due date"]] or "").strip() if "Due date" in idx and row[idx["Due date"]] is not None else ""
        document = str(row[idx["เอกสาร"]] or "").strip() if "เอกสาร" in idx and row[idx["เอกสาร"]] is not None else ""
        detail = str(row[idx["Detail (รายละเอียดข้อหา)"]] or "").strip() if "Detail (รายละเอียดข้อหา)" in idx and row[idx["Detail (รายละเอียดข้อหา)"]] is not None else ""
        
        if not cid and not name:
            continue

        record = {
            "id": cid,
            "name": name,
            "status": status,
            "area": area,
            "caught_area": caught_area,
            "occurance_date": occurance_date,
            "company_name": company_name,
            "department": department,
            "card_number": card_number,
            "position": position,
            "affiliation": affiliation,
            "allegation": allegation,
            "start_date": start_date,
            "due_date": due_date,
            "document": document,
            "detail": detail,
        }

        if cid:
            id_records[cid] = record

        if name:
            name_records[name] = record

    return id_records, name_records


# Blacklist check
def check_blacklist(card_data, id_records, name_records):
    cid = clean_id(card_data.get("CID", ""))
    nameTH = normalize(card_data.get("TH Fullname", ""))
    nameEN = normalize(card_data.get("EN Fullname", ""))

    def create_result(reason_text, record):
        return True, {
            "reason": reason_text,
            "status": record["status"],
            "area": record["area"],
            "caught_area": record["caught_area"],
            "occurance_date": record["occurance_date"],
            "company_name": record["company_name"],
            "department": record["department"],
            "card_number": record["card_number"],
            "position": record["position"],
            "affiliation": record["affiliation"],
            "allegation": record["allegation"],
            "start_date": record["start_date"],
            "due_date": record["due_date"],
            "document": record["document"],
            "detail": record["detail"],
        }

    # Match by ID
    if cid and cid in id_records:
        return create_result(f"พบเลขบัตรเหมือนกัน -> {cid}", id_records[cid])

    # Match by Thai name
    if nameTH and nameTH in name_records:
        return create_result(f"พบชื่อ-นามสกุลเหมือนกัน -> {nameTH}", name_records[nameTH])

    # Match by English name
    if nameEN and nameEN in name_records:
        return create_result(f"พบชื่อ-นามสกุลเหมือนกัน -> {nameEN}", name_records[nameEN])

    return False, {
        "reason": "",
        "status": "",
        "area": "",
        "caught_area": "",
        "occurance_date": "",
        "company_name": "",
        "department": "",
        "card_number": "",
        "position": "",
        "affiliation": "",
        "allegation": "",
        "start_date": "",
        "due_date": "",
        "document": "",
        "detail": "",
    }


# APDU command bytes  
APDU_SELECT = [0x00, 0xA4, 0x04, 0x00, 0x08, 0xA0, 0x00, 0x00, 0x00, 0x54, 0x48, 0x00, 0x01]
FIELD = [
    ("CID",             [0x80, 0xB0, 0x00, 0x04, 0x02, 0x00, 0x0D]),
    ("TH Fullname",     [0x80, 0xB0, 0x00, 0x11, 0x02, 0x00, 0x64]),
    ("EN Fullname",     [0x80, 0xB0, 0x00, 0x75, 0x02, 0x00, 0x64]),
    ("Date of Birth",   [0x80, 0xB0, 0x00, 0xD9, 0x02, 0x00, 0x08]),
    ("Gender",          [0x80, 0xB0, 0x00, 0xE1, 0x02, 0x00, 0x01]),
    ("Card Issuer",     [0x80, 0xB0, 0x00, 0xF6, 0x02, 0x00, 0x64]),
    ("Issue Date",      [0x80, 0xB0, 0x01, 0x67, 0x02, 0x00, 0x08]),
    ("Expire Date",     [0x80, 0xB0, 0x01, 0x6F, 0x02, 0x00, 0x08]),
    ("Address",         [0x80, 0xB0, 0x15, 0x79, 0x02, 0x00, 0x64]),
]


# Card reading
def transmit_APDU(conn, apdu):
    data, sw1, sw2 = conn.transmit(apdu)
    if sw1 == 0x61:
        data, sw1, sw2 = conn.transmit([0x00, 0xC0, 0x00, 0x00, sw2])
    return data, sw1, sw2


def thai2unicode(data):
    return bytes(data).decode("tis-620", errors="replace").replace("#", " ").strip()


def read_card(reader_list):
    card_data = {}
    try:
        with reader_list[0].createConnection() as connection:
            connection.connect()
            data, sw1, sw2 = transmit_APDU(connection, APDU_SELECT)
            if sw1 != 0x90:
                print("Applet selection failed")
                return card_data
            for (label, cmd) in FIELD:
                data, sw1, sw2 = transmit_APDU(connection, cmd)
                if sw1 == 0x90 and data:
                    value = thai2unicode(data)
                    if label == "Gender":
                        value = {"1": "ชาย", "2": "หญิง"}.get(value, value)
                    card_data[label] = value
    except NoCardException:
        pass
    return card_data


# Terminal display
SEPARATOR = "=" * 50
def print_card(card_data: dict):
    print(SEPARATOR)
    for label, value in card_data.items():
        print(f"  {label:15}: {value}")


def print_result(is_match: bool, result: dict):
    print(SEPARATOR)

    if is_match:
        print(result["status"].upper())
        print(f"Reason      : {result['reason']}")
        print(f"Status      : {result['status']}")
        print(f"Allegation  : {result['allegation']}")
        print(f"Start Date  : {result['start_date']}")
        print(f"Due Date    : {result['due_date']}")
    else:
        print("CLEAR")

    print(SEPARATOR)


def main_polling(id_records, name_records, callback=None, stop_event=None):
    """
    Polls card readers and updates card status via callback.
    callback signature: callback(event_type, data)
    event_type: "NO_READER", "WAITING_CARD", "NO_CARD", "CARD_READ"
    """
    last_cid = None
    last_state = "unknown"

    while stop_event is None or not stop_event.is_set():
        try:
            reader_list = readers()
            # 1. No reader connected
            if not reader_list:
                if last_state != "no_reader":
                    last_state = "no_reader"
                    last_cid = None
                    if callback:
                        callback("NO_READER", None)
                    else:
                        print("No card reader detected.")
                time.sleep(1)
                continue
            # 2. Reader connected
            if last_state != "has_reader":
                last_state = "has_reader"
                last_cid = None
                if callback:
                    callback("WAITING_CARD", None)
                else:
                    print("Waiting for card...\n")
            # 3. Read card
            card_data = read_card(reader_list)
            if card_data:
                cid = card_data.get("CID", "")
                if cid != last_cid: # New card inserted
                    last_cid = cid
                    is_bl, result = check_blacklist(card_data, id_records, name_records)
                    
                    if callback:
                        callback("CARD_READ", {
                            "card_data": card_data,
                            "is_match": is_bl,
                            "result": result
                        })
                    else:
                        print_card(card_data)
                        print_result(is_bl, result)
            else:
                if last_cid is not None: # Card removed
                    last_cid = None
                    if callback:
                        callback("NO_CARD", None)
                    else:
                        print("No card inserted!")
        except Exception as e:
            last_cid = None
        time.sleep(1)
